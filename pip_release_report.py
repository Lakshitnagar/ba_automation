#!/usr/bin/env python3
import argparse
import csv
import datetime as dt
import json
import re
from pathlib import Path

import requests

try:
    from packaging.version import parse as vparse
except Exception:  # pragma: no cover - best-effort import
    vparse = None

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
except Exception as exc:  # pragma: no cover - best-effort import
    raise SystemExit(
        "openpyxl is required to write the Excel file. "
        "Install it with: pip install openpyxl"
    ) from exc

LINE_RE = re.compile(r"^\s*([A-Za-z0-9_.-]+)\s*==\s*([^\s;]+)")
NPM_VERSION_RE = re.compile(r"(\d+\.\d+\.\d+(?:[-+][0-9A-Za-z.-]+)?)")
PYPI_URL = "https://pypi.org/pypi/{name}/json"
NPM_URL = "https://registry.npmjs.org/{name}"
BA_LIST_PATH = "ba_list.csv"
BA_LINK_TEMPLATE = (
    "https://pls.appoci.oraclecorp.com/PLS/faces/ThirdPartyHome?wid={ba_id}"
)
EXCLUDED_PACKAGES = {
    "colorlog",
    "django-extensions",
    "redis",
    "setuptools",
    "wheel",
}


def parse_pip_file(path: Path) -> list[tuple[str, str]]:
    items: list[tuple[str, str]] = []
    for raw in path.read_text(encoding="utf-8").splitlines():
        line = raw.strip()
        if not line or line.startswith("#"):
            continue
        match = LINE_RE.match(line)
        if not match:
            continue
        name, version = match.group(1), match.group(2)
        items.append((name, version))
    return items


def parse_package_json(path: Path) -> list[tuple[str, str]]:
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return []
    items: list[tuple[str, str]] = []
    deps = data.get("dependencies", {})
    if not isinstance(deps, dict):
        return items
    for name, spec in deps.items():
        if not isinstance(spec, str):
            continue
        if name.startswith("@angular/"):
            continue
        items.append((name, spec))
    return items


def extract_npm_version(spec: str) -> str | None:
    match = NPM_VERSION_RE.search(spec)
    return match.group(1) if match else None


def load_ba_map(
    path: Path,
) -> dict[tuple[str, str], dict[str, dict[str, dt.date | str | None]]]:
    if not path.exists():
        return {}
    mapping: dict[tuple[str, str], dict[str, dict[str, dt.date | str | None]]] = {}
    with path.open(encoding="utf-8", errors="replace") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            name = (row.get("Licensed Item Name") or "").strip()
            version = (row.get("Licensed Item Version") or "").strip()
            ba_id = (row.get("Business Approval ID") or "").strip()
            created = (row.get("Created Date") or "").strip()
            ba_end_date = (row.get("BA End Date") or "").strip()
            ba_end_action = (row.get("BA End Date Action") or "").strip()
            ba_status = (row.get("Status") or "").strip()
            if not name or not version or not ba_id:
                continue
            created_value: dt.date | str | None = None
            if created:
                try:
                    created_value = dt.date.fromisoformat(created)
                except ValueError:
                    created_value = created
            end_date_value: dt.date | str | None = None
            if ba_end_date:
                try:
                    end_date_value = dt.date.fromisoformat(ba_end_date)
                except ValueError:
                    end_date_value = ba_end_date
            key = (name.lower(), version)
            mapping.setdefault(key, {})
            if ba_id not in mapping[key]:
                mapping[key][ba_id] = {
                    "created": created_value,
                    "end_date": end_date_value,
                    "end_action": ba_end_action or None,
                    "status": ba_status or None,
                }
    return mapping


def get_release_date(releases: dict, version: str) -> dt.date | None:
    files = releases.get(version) or []
    dates = []
    for entry in files:
        ts = entry.get("upload_time_iso_8601") or entry.get("upload_time")
        if not ts:
            continue
        try:
            dates.append(dt.datetime.fromisoformat(ts.replace("Z", "+00:00")))
        except ValueError:
            continue
    if not dates:
        return None
    return min(dates).date()


def is_stable_version(version: str) -> bool:
    if not vparse:
        return False
    parsed = vparse(version)
    # Exclude pre, post, dev, and local versions
    return not (parsed.is_prerelease or parsed.is_postrelease or parsed.is_devrelease or parsed.local)


def get_latest_version(info: dict, releases: dict) -> str | None:
    if not vparse:
        return info.get("version")
    versions = [v for v in releases.keys() if v and is_stable_version(v)]
    if not versions:
        return None
    return str(max(versions, key=vparse))


def get_latest_version_same_major(
    releases: dict, current_version: str
) -> str | None:
    if not vparse:
        return None
    try:
        current_major = vparse(current_version).release[0]
    except Exception:
        return None
    candidates = []
    for version in releases.keys():
        if not version or not is_stable_version(version):
            continue
        parsed = vparse(version)
        if parsed.release and parsed.release[0] == current_major:
            candidates.append(version)
    if not candidates:
        return None
    return str(max(candidates, key=vparse))


def sanitize_sheet_name(name: str) -> str:
    cleaned = re.sub(r"[\\/*?:\\[\\]]", "_", name)
    return cleaned[:31] if cleaned else "Sheet"


def fetch_pypi(name: str, session: requests.Session, cache: dict) -> dict | None:
    if name in cache:
        return cache[name]
    try:
        resp = session.get(PYPI_URL.format(name=name), timeout=20)
        if resp.status_code != 200:
            cache[name] = None
            return None
        cache[name] = resp.json()
        return cache[name]
    except requests.RequestException:
        cache[name] = None
        return None


def fetch_npm(name: str, session: requests.Session, cache: dict) -> dict | None:
    if name in cache:
        return cache[name]
    try:
        encoded = requests.utils.quote(name, safe="@/")
        resp = session.get(NPM_URL.format(name=encoded), timeout=20)
        if resp.status_code != 200:
            cache[name] = None
            return None
        cache[name] = resp.json()
        return cache[name]
    except requests.RequestException:
        cache[name] = None
        return None


def get_npm_release_date(time_map: dict, version: str) -> dt.date | None:
    if not version:
        return None
    ts = time_map.get(version)
    if not ts:
        return None
    try:
        return dt.datetime.fromisoformat(ts.replace("Z", "+00:00")).date()
    except ValueError:
        return None


def get_npm_latest_version(data: dict) -> str | None:
    dist_tags = data.get("dist-tags", {}) if isinstance(data, dict) else {}
    latest = dist_tags.get("latest")
    if not vparse:
        return latest if isinstance(latest, str) else None
    if isinstance(latest, str) and is_stable_version(latest):
        return latest
    time_map = data.get("time", {}) if isinstance(data, dict) else {}
    versions = [
        v
        for v in time_map.keys()
        if v not in ("created", "modified") and is_stable_version(v)
    ]
    if not versions:
        return None
    return str(max(versions, key=vparse))


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Generate an Excel report comparing pinned versions to latest PyPI/NPM releases."
    )
    parser.add_argument(
        "--root",
        default=".",
        help="Root directory to scan for .pip files (default: current directory).",
    )
    parser.add_argument(
        "--output",
        default="pip_release_report.xlsx",
        help="Output Excel file path (default: pip_release_report.xlsx).",
    )
    args = parser.parse_args()

    root = Path(args.root).resolve()
    pip_files = list(root.rglob("*.pip"))
    npm_files = list(root.rglob("package.json"))
    if not pip_files and not npm_files:
        print(f"No .pip or package.json files found under {root}")
        return 1

    grouped: dict[str, list[tuple[str, str, str]]] = {}
    for path in pip_files:
        folder = path.parent.name or path.parent.as_posix()
        grouped.setdefault(folder, [])
        grouped[folder].extend((name, version, "pypi") for name, version in parse_pip_file(path))
    for path in npm_files:
        folder = path.parent.name or path.parent.as_posix()
        grouped.setdefault(folder, [])
        grouped[folder].extend((name, version, "npm") for name, version in parse_package_json(path))

    ba_map = load_ba_map(root / BA_LIST_PATH)

    wb = Workbook()
    wb.remove(wb.active)

    session = requests.Session()
    session.headers.update({"User-Agent": "pip-release-report/1.0"})
    pypi_cache: dict[str, dict | None] = {}
    npm_cache: dict[str, dict | None] = {}

    headers = [
        "package",
        "current_version",
        "current_release_date",
        "latest_version",
        "latest_release_date",
        "days_difference",
        "days_since_latest_release",
        "days_since_current_release",
        "business_approval_ids",
        "business_approval_status",
        "business_approval_created_date",
        "business_approval_end_date",
        "business_approval_end_date_action",
    ]
    summary_sections = ("sources", "tipcms", "collection")
    flagged_by_section: dict[str, list[list]] = {}
    zero_diff_by_section: dict[str, list[list]] = {}
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True, size=16)
    link_font = Font(color="0563C1", underline="single")
    header_alignment = Alignment(horizontal="center", vertical="center")
    body_font = Font(size=14)
    body_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    package_alignment = Alignment(horizontal="left", vertical="center")
    even_fill = PatternFill("solid", fgColor="F2F6FA")
    odd_fill = PatternFill("solid", fgColor="FFFFFF")
    alert_fill = PatternFill("solid", fgColor="F8D7DA")
    warning_fill = PatternFill("solid", fgColor="FFF3CD")
    ba_ok_fill = PatternFill("solid", fgColor="D4EDDA")
    ba_bad_fill = PatternFill("solid", fgColor="F8D7DA")
    alert_threshold_days = 2 * 365 - 2 * 31  # ~2 years minus ~2 months

    today = dt.date.today()
    for folder, items in sorted(grouped.items()):
        ws = wb.create_sheet(title=sanitize_sheet_name(folder))
        ws.append(headers)
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        rows_info: list[
            tuple[list, bool, list[tuple[str, dict[str, dt.date | str | None]]]]
        ] = []
        for name, current_version, ecosystem in items:
            if name.lower() in EXCLUDED_PACKAGES:
                continue
            if ecosystem == "pypi":
                data = fetch_pypi(name, session, pypi_cache)
                if not data:
                    rows_info.append(
                        (
                            [
                                name,
                                current_version,
                                None,
                                None,
                                None,
                                None,
                                None,
                                None,
                                None,
                                None,
                                None,
                                None,
                                None,
                            ],
                            False,
                        )
                    )
                    continue

                info = data.get("info", {})
                releases = data.get("releases", {})
                current_date = get_release_date(releases, current_version)
                if name.lower() == "django":
                    latest_version = get_latest_version_same_major(releases, current_version)
                    if not latest_version:
                        latest_version = get_latest_version(info, releases)
                else:
                    latest_version = get_latest_version(info, releases)
                latest_date = get_release_date(releases, latest_version) if latest_version else None
                ba_version = current_version
            else:
                data = fetch_npm(name, session, npm_cache)
                if not data:
                    rows_info.append(
                        (
                            [
                                name,
                                current_version,
                                None,
                                None,
                                None,
                                None,
                                None,
                                None,
                                None,
                                None,
                                None,
                                None,
                                None,
                            ],
                            False,
                        )
                    )
                    continue

                time_map = data.get("time", {})
                current_resolved = extract_npm_version(current_version)
                current_date = get_npm_release_date(time_map, current_resolved)
                latest_version = get_npm_latest_version(data)
                latest_date = get_npm_release_date(time_map, latest_version)
                ba_version = current_resolved

            ba_entries: list[tuple[str, dict[str, dt.date | str | None]]] = []
            if ba_version:
                ba_ids_map = ba_map.get((name.lower(), ba_version))
                if ba_ids_map:
                    def ba_sort_key(item: tuple[str, dict[str, dt.date | str | None]]):
                        ba_id, meta = item
                        end_date = meta.get("end_date")
                        if isinstance(end_date, dt.date):
                            key_date = end_date
                        elif isinstance(end_date, str):
                            try:
                                key_date = dt.date.fromisoformat(end_date)
                            except ValueError:
                                key_date = None
                        else:
                            key_date = None
                        status = (meta.get("status") or "").lower()
                        status_priority = 1 if status == "approved" else 0
                        return (key_date is None, key_date or dt.date.max, status_priority, ba_id)

                    ba_entries = sorted(ba_ids_map.items(), key=ba_sort_key, reverse=True)

            days_diff = None
            if current_date and latest_date:
                days_diff = (latest_date - current_date).days
            days_since_latest = None
            if latest_date:
                days_since_latest = (today - latest_date).days
            days_since_current = None
            if current_date:
                days_since_current = (today - current_date).days

            row = [
                name,
                current_version,
                current_date if current_date else None,
                latest_version,
                latest_date if latest_date else None,
                days_diff,
                days_since_latest,
                days_since_current,
                None,
                None,
                None,
                None,
                None,
            ]
            is_alert = (
                isinstance(days_since_current, int)
                and days_since_current > alert_threshold_days
                and isinstance(days_diff, int)
                and days_diff > 0
            )
            rows_info.append((row, is_alert, ba_entries))
        rows_info.sort(
            key=lambda r: (r[0][7] is None, r[0][7] if r[0][7] is not None else -1),
            reverse=True,
        )
        ba_blocks: list[tuple[int, int, PatternFill]] = []
        for row, is_alert, ba_entries in rows_info:
            start_row = ws.max_row + 1
            expanded_rows: list[list] = []
            ba_fill = None
            if ba_entries:
                top_status = (ba_entries[0][1].get("status") or "").strip().lower()
                ba_fill = ba_ok_fill if top_status == "approved" else ba_bad_fill
            if not ba_entries:
                ws.append(row)
                expanded_rows.append(row)
                end_row = start_row
            else:
                for ba_id, ba_meta in ba_entries:
                    row_with_ba = row.copy()
                    row_with_ba[8] = ba_id
                    row_with_ba[9] = ba_meta.get("status")
                    row_with_ba[10] = ba_meta.get("created")
                    row_with_ba[11] = ba_meta.get("end_date")
                    row_with_ba[12] = ba_meta.get("end_action")
                    ws.append(row_with_ba)
                    ba_cell = ws.cell(row=ws.max_row, column=9)
                    ba_cell.hyperlink = BA_LINK_TEMPLATE.format(ba_id=ba_id)
                    ba_cell.font = link_font
                    expanded_rows.append(row_with_ba)
                end_row = start_row + len(ba_entries) - 1
                if end_row > start_row:
                    for col in range(1, len(headers) + 1):
                        if col in (9, 10, 11, 12, 13):
                            continue
                        ws.merge_cells(
                            start_row=start_row,
                            start_column=col,
                            end_row=end_row,
                            end_column=col,
                        )
            if ba_fill:
                ba_blocks.append((start_row, start_row, ba_fill))
            if is_alert and folder in summary_sections:
                flagged_by_section.setdefault(folder, []).extend(expanded_rows)
            if (
                folder in summary_sections
                and row[5] == 0
                and isinstance(row[6], int)
                and row[6] > alert_threshold_days
            ):
                zero_diff_by_section.setdefault(folder, []).extend(expanded_rows)
        for row in range(2, ws.max_row + 1):
            days_value = ws.cell(row=row, column=8).value
            diff_value = ws.cell(row=row, column=6).value
            if (
                isinstance(days_value, int)
                and days_value > alert_threshold_days
                and isinstance(diff_value, int)
                and diff_value > 0
            ):
                fill = alert_fill
            else:
                fill = even_fill if row % 2 == 0 else odd_fill
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=row, column=col)
                cell.fill = fill
                cell.font = body_font
                cell.alignment = package_alignment if col == 1 else body_alignment
                if col in (3, 5, 11, 12) and cell.value:
                    cell.number_format = "DD-MMM-YYYY"
                if col == 7 and isinstance(cell.value, int) and cell.value > alert_threshold_days:
                    cell.fill = warning_fill
                if col in (9, 10, 11, 12, 13):
                    if not ws.cell(row=row, column=9).value:
                        cell.fill = ba_bad_fill
                    else:
                        for start_row, end_row, ba_fill in ba_blocks:
                            if start_row <= row <= end_row:
                                cell.fill = ba_fill
                                break
                if col == 12 and isinstance(cell.value, dt.date):
                    days_until_end = (cell.value - today).days
                    if 0 <= days_until_end <= 90:
                        cell.fill = alert_fill
        for col in range(1, len(headers) + 1):
            max_len = 0
            for row in range(1, ws.max_row + 1):
                value = ws.cell(row=row, column=col).value
                if value is None:
                    continue
                max_len = max(max_len, len(str(value)))
            header_len = len(str(headers[col - 1]))
            width = max(max_len + 2, int(header_len * 1.25) + 4)
            ws.column_dimensions[chr(64 + col)].width = width

    summary_headers = ["section", *headers]
    summary_ws = wb.create_sheet(title="Upgradation")
    summary_ws.append(summary_headers)
    for col in range(1, len(summary_headers) + 1):
        cell = summary_ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    summary_rows: list[list] = []
    for section in summary_sections:
        for row in flagged_by_section.get(section, []):
            summary_rows.append([section, *row])
    summary_rows.sort(
        key=lambda r: (r[8] is None, r[8] if r[8] is not None else -1),
        reverse=True,
    )
    for row in summary_rows:
        summary_ws.append(row)
        ba_id = summary_ws.cell(row=summary_ws.max_row, column=10).value
        if ba_id:
            ba_cell = summary_ws.cell(row=summary_ws.max_row, column=10)
            ba_cell.hyperlink = BA_LINK_TEMPLATE.format(ba_id=ba_id)
            ba_cell.font = link_font

    for row in range(2, summary_ws.max_row + 1):
        for col in range(1, len(summary_headers) + 1):
            cell = summary_ws.cell(row=row, column=col)
            cell.fill = alert_fill
            cell.font = body_font
            cell.alignment = package_alignment if col == 2 else body_alignment
            if col in (4, 6, 11, 12) and cell.value:
                cell.number_format = "DD-MMM-YYYY"
            if col == 8 and isinstance(cell.value, int) and cell.value > alert_threshold_days:
                cell.fill = warning_fill

    for col in range(1, len(summary_headers) + 1):
        max_len = 0
        for row in range(1, summary_ws.max_row + 1):
            value = summary_ws.cell(row=row, column=col).value
            if value is None:
                continue
            max_len = max(max_len, len(str(value)))
        header_len = len(str(summary_headers[col - 1]))
        width = max(max_len + 2, int(header_len * 1.25) + 4)
        summary_ws.column_dimensions[chr(64 + col)].width = width
    if summary_ws.max_row > 2:
        non_ba_last_col = len(summary_headers) - 5
        start_row = 2
        while start_row <= summary_ws.max_row:
            end_row = start_row
            while end_row + 1 <= summary_ws.max_row:
                if all(
                    summary_ws.cell(row=end_row + 1, column=col).value
                    == summary_ws.cell(row=start_row, column=col).value
                    for col in range(1, non_ba_last_col + 1)
                ):
                    end_row += 1
                else:
                    break
            if end_row > start_row:
                for col in range(1, non_ba_last_col + 1):
                    summary_ws.merge_cells(
                        start_row=start_row,
                        start_column=col,
                        end_row=end_row,
                        end_column=col,
                    )
            start_row = end_row + 1
    if summary_ws.max_row > 2:
        non_ba_last_col = len(summary_headers) - 5
        start_row = 2
        while start_row <= summary_ws.max_row:
            end_row = start_row
            while end_row + 1 <= summary_ws.max_row:
                if all(
                    summary_ws.cell(row=end_row + 1, column=col).value
                    == summary_ws.cell(row=start_row, column=col).value
                    for col in range(1, non_ba_last_col + 1)
                ):
                    end_row += 1
                else:
                    break
            ba_id_val = (summary_ws.cell(row=start_row, column=10).value or "").strip()
            if ba_id_val:
                status_val = (
                    summary_ws.cell(row=start_row, column=11).value or ""
                ).strip().lower()
                ba_fill = ba_ok_fill if status_val == "approved" else ba_bad_fill
            else:
                ba_fill = ba_bad_fill
            for col in range(10, 15):
                summary_ws.cell(row=start_row, column=col).fill = ba_fill
            start_row = end_row + 1
        for row in range(2, summary_ws.max_row + 1):
            end_date_val = summary_ws.cell(row=row, column=13).value
            if isinstance(end_date_val, dt.date):
                days_until_end = (end_date_val - today).days
                if 0 <= days_until_end <= 90:
                    summary_ws.cell(row=row, column=13).fill = alert_fill

    zero_ws = wb.create_sheet(title="Replace-Remove Libs")
    zero_ws.append(summary_headers)
    for col in range(1, len(summary_headers) + 1):
        cell = zero_ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    zero_rows: list[list] = []
    for section in summary_sections:
        for row in zero_diff_by_section.get(section, []):
            zero_rows.append([section, *row])
    zero_rows.sort(
        key=lambda r: (r[8] is None, r[8] if r[8] is not None else -1),
        reverse=True,
    )
    for row in zero_rows:
        zero_ws.append(row)
        ba_id = zero_ws.cell(row=zero_ws.max_row, column=10).value
        if ba_id:
            ba_cell = zero_ws.cell(row=zero_ws.max_row, column=10)
            ba_cell.hyperlink = BA_LINK_TEMPLATE.format(ba_id=ba_id)
            ba_cell.font = link_font

    for row in range(2, zero_ws.max_row + 1):
        days_latest = zero_ws.cell(row=row, column=8).value
        row_fill = (
            warning_fill
            if isinstance(days_latest, int) and days_latest > alert_threshold_days
            else alert_fill
        )
        for col in range(1, len(summary_headers) + 1):
            cell = zero_ws.cell(row=row, column=col)
            cell.fill = row_fill
            cell.font = body_font
            cell.alignment = package_alignment if col == 2 else body_alignment
            if col in (4, 6, 11, 12) and cell.value:
                cell.number_format = "DD-MMM-YYYY"

    for col in range(1, len(summary_headers) + 1):
        max_len = 0
        for row in range(1, zero_ws.max_row + 1):
            value = zero_ws.cell(row=row, column=col).value
            if value is None:
                continue
            max_len = max(max_len, len(str(value)))
        header_len = len(str(summary_headers[col - 1]))
        width = max(max_len + 2, int(header_len * 1.25) + 4)
        zero_ws.column_dimensions[chr(64 + col)].width = width
    if zero_ws.max_row > 2:
        non_ba_last_col = len(summary_headers) - 5
        start_row = 2
        while start_row <= zero_ws.max_row:
            end_row = start_row
            while end_row + 1 <= zero_ws.max_row:
                if all(
                    zero_ws.cell(row=end_row + 1, column=col).value
                    == zero_ws.cell(row=start_row, column=col).value
                    for col in range(1, non_ba_last_col + 1)
                ):
                    end_row += 1
                else:
                    break
            if end_row > start_row:
                for col in range(1, non_ba_last_col + 1):
                    zero_ws.merge_cells(
                        start_row=start_row,
                        start_column=col,
                        end_row=end_row,
                        end_column=col,
                    )
            start_row = end_row + 1
    if zero_ws.max_row > 2:
        non_ba_last_col = len(summary_headers) - 5
        start_row = 2
        while start_row <= zero_ws.max_row:
            end_row = start_row
            while end_row + 1 <= zero_ws.max_row:
                if all(
                    zero_ws.cell(row=end_row + 1, column=col).value
                    == zero_ws.cell(row=start_row, column=col).value
                    for col in range(1, non_ba_last_col + 1)
                ):
                    end_row += 1
                else:
                    break
            ba_id_val = (zero_ws.cell(row=start_row, column=10).value or "").strip()
            if ba_id_val:
                status_val = (zero_ws.cell(row=start_row, column=11).value or "").strip().lower()
                ba_fill = ba_ok_fill if status_val == "approved" else ba_bad_fill
            else:
                ba_fill = ba_bad_fill
            for col in range(10, 15):
                zero_ws.cell(row=start_row, column=col).fill = ba_fill
            start_row = end_row + 1
        for row in range(2, zero_ws.max_row + 1):
            end_date_val = zero_ws.cell(row=row, column=13).value
            if isinstance(end_date_val, dt.date):
                days_until_end = (end_date_val - today).days
                if 0 <= days_until_end <= 90:
                    zero_ws.cell(row=row, column=13).fill = alert_fill

    desired_order = [
        "tipcms",
        "sources",
        "collection",
        "Upgradation",
        "Replace-Remove Libs",
    ]
    name_to_sheet = {sheet.title: sheet for sheet in wb.worksheets}
    wb._sheets = [name_to_sheet[name] for name in desired_order if name in name_to_sheet]
    wb._sheets.extend(
        sheet for sheet in wb.worksheets if sheet.title not in desired_order
    )

    wb.save(args.output)
    print(f"Wrote {args.output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
